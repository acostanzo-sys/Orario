import os
import openpyxl
from datetime import datetime
from flask import current_app


# ============================================================
#   DIAGNOSTICA SOVRAPPOSIZIONI (XLS FINALE)
# ============================================================

def diagnostica_sovrapposizioni(calendario_per_classe):
    conflitti = []
    occupazione = {}  # occupazione[docente][data][ora] = [classi]

    for classe_id, dati in calendario_per_classe.items():
        nome_classe = dati["nome_classe"]

        for giorno in dati["calendario"]:
            data = giorno["data"]
            lezioni = giorno["lezioni"]

            for idx, slot in enumerate(lezioni):
                docente = slot["docente"]
                if not docente:
                    continue

                if docente not in occupazione:
                    occupazione[docente] = {}
                if data not in occupazione[docente]:
                    occupazione[docente][data] = {}
                if idx not in occupazione[docente][data]:
                    occupazione[docente][data][idx] = []

                occupazione[docente][data][idx].append(nome_classe)

    for docente, giorni in occupazione.items():

        if docente.strip().upper() == "DOC EST":
            continue

        for data, ore in giorni.items():
            for ora, classi in ore.items():
                if len(classi) > 1:
                    conflitti.append({
                        "docente": docente,
                        "data": data,
                        "ora": ora,
                        "classi": classi
                    })

    return conflitti



def diagnostica_ultimo_calendario():
    folder = os.path.join(current_app.root_path, "generated_calendars")
    files = sorted(os.listdir(folder), reverse=True)

    if not files:
        return [{"errore": "Nessun calendario generato."}]

    ultimo = files[0]
    path = os.path.join(folder, ultimo)

    wb = openpyxl.load_workbook(path)
    calendario_per_classe = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        calendario = []
        current_day = None
        lezioni_giorno = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            data_str, giorno_label, ora_str, materia, docente = row

            if not data_str:
                continue

            data = datetime.strptime(data_str, "%d/%m/%Y").date()

            if current_day != data:
                if current_day is not None:
                    calendario.append({
                        "data": current_day,
                        "giorno_settimana": giorno_label_prev,
                        "lezioni": lezioni_giorno
                    })
                current_day = data
                giorno_label_prev = giorno_label
                lezioni_giorno = []

            ora = datetime.strptime(ora_str, "%H:%M").time()
            lezioni_giorno.append({
                "ora": ora,
                "materia": materia,
                "docente": docente
            })

        if current_day is not None:
            calendario.append({
                "data": current_day,
                "giorno_settimana": giorno_label_prev,
                "lezioni": lezioni_giorno
            })

        calendario_per_classe[sheet] = {
            "nome_classe": sheet,
            "ore_giornaliere": 6,
            "calendario": calendario
        }

    return diagnostica_sovrapposizioni(calendario_per_classe)



# ============================================================
#   DIAGNOSTICA PER SINGOLA CLASSE (slot bloccati)
# ============================================================

def diagnostica_classe(cd, docente_ok_wrapper):
    print("\n\n======================")
    print(f" DIAGNOSTICA CLASSE {cd['classe'].nome_classe}")
    print("======================")

    materie = cd["materie_attive"]
    griglie = cd["griglie"]
    ore_g = cd["ore_g"]
    tipo_giorno = cd["tipo_giorno"]

    print("\n--- REPORT MATERIE ---")
    for mid, info in materie.items():
        nome = info["nome"]
        ore_ann = info.get("ore_annuali", 0)
        assegnate = info.get("ore_assegnate", 0)
        debito = info.get("debito_residuo", 0)

        print(f"\nMateria: {nome} (ID {mid})")
        print(f"  Ore annue:      {ore_ann}")
        print(f"  Ore piazzate:   {assegnate}")
        print(f"  Debito residuo: {debito}")

        if debito > 0:
            print("  ⚠ Mancano ore! Analizzo gli slot bloccati...")

            for key, giorni in cd["giorni_per_key"].items():
                griglia = griglie[key]
                for g in giorni:
                    data_g = g["data"]
                    giorno_it = g["giorno_it"]
                    row = griglia[data_g]

                    for h in range(ore_g):
                        slot = row[h]

                        if slot is None:
                            ok_doc = docente_ok_wrapper(info.get("docente_id"), data_g, h, giorno_it, 1)

                            intoccabile = False
                            if (data_g, h) in cd.get("fissi_per_giorno", set()):
                                intoccabile = True
                            if (data_g, h) in cd.get("speciali_per_giorno", set()):
                                intoccabile = True
                            if tipo_giorno.get(data_g) in ("STAGE", "FESTA", "SPECIALE"):
                                intoccabile = True

                            if intoccabile:
                                print(f"    - {data_g} ora {h}: ❌ slot intoccabile")
                            elif not ok_doc:
                                print(f"    - {data_g} ora {h}: ❌ docente non disponibile")
                            else:
                                print(f"    - {data_g} ora {h}: ✔ slot libero e piazzabile (ma non usato)")

    print("\n--- FINE DIAGNOSTICA ---\n")



# ============================================================
#   DIAGNOSTICA GLOBALE
# ============================================================

def diagnostica_globale(classi_data, docente_ok_wrapper):
    print("\n\n=======================================")
    print("        DIAGNOSTICA GLOBALE ORARIO")
    print("=======================================\n")

    totale_ore_attese = 0
    totale_ore_piazzate = 0
    totale_ore_residue = 0

    for cd in classi_data:
        classe = cd["classe"].nome_classe
        materie = cd["materie_attive"]

        print(f"\n\n===== CLASSE {classe} =====")

        ore_attese_classe = 0
        ore_piazzate_classe = 0
        ore_residue_classe = 0

        for mid, info in materie.items():
            nome = info["nome"]
            ore_ann = info.get("ore_annuali", 0)
            assegnate = info.get("ore_assegnate", 0)
            debito = info.get("debito_residuo", 0)

            ore_attese_classe += ore_ann
            ore_piazzate_classe += assegnate
            ore_residue_classe += debito

            if debito > 0:
                print(f"  ⚠ {nome}: mancano {debito} ore")

        print(f"\n  → Totale ore previste: {ore_attese_classe}")
        print(f"  → Ore piazzate:        {ore_piazzate_classe}")
        print(f"  → Ore residue:         {ore_residue_classe}")

        totale_ore_attese += ore_attese_classe
        totale_ore_piazzate += ore_piazzate_classe
        totale_ore_residue += ore_residue_classe

    print("\n\n=======================================")
    print("           RIEPILOGO FINALE")
    print("=======================================\n")
    print(f"Ore totali previste: {totale_ore_attese}")
    print(f"Ore totali piazzate: {totale_ore_piazzate}")
    print(f"Ore totali residue:  {totale_ore_residue}")

    if totale_ore_residue == 0:
        print("\n✔ TUTTE LE ORE SONO STATE PIAZZATE!")
    else:
        print("\n⚠ CI SONO ORE NON PIAZZATE — VEDI DETTAGLI SOPRA")
