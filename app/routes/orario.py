# app/routes/orario.py

from flask import Blueprint, render_template, send_file, redirect, request, url_for, flash, current_app
from datetime import datetime
from io import BytesIO
import os
import openpyxl
from flask import send_from_directory

from app.models import (
    db,
    Classe,
    MateriaClasse,
    Materia,
    AnnoFormativo,
    GiornoFisso,
    VincoloDocente,
    GiornoSpeciale,
    Stage,
    Docente,
)

import app.utils.orario_utils as orario_utils
from app.utils.calendario_generator import genera_calendario_annuale

orario_bp = Blueprint("orario", __name__, url_prefix="/orario")

REPORT_VINCOLI = {}

# ============================================================
# 1) GENERA CALENDARIO (POST)
# ============================================================

@orario_bp.route("/genera", methods=["POST"])
def genera_calendario():
    calendario = genera_calendario_annuale()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for classe_id, dati in calendario.items():
        ws = wb.create_sheet(title=dati["nome_classe"][:31])

        ws.cell(row=1, column=1, value="Data")
        ws.cell(row=1, column=2, value="Giorno")
        ws.cell(row=1, column=3, value="Ora")
        ws.cell(row=1, column=4, value="Materia")
        ws.cell(row=1, column=5, value="Docente")

        row = 2
        for giorno in dati["calendario"]:
            data_str = giorno["data"].strftime("%d/%m/%Y")
            giorno_label = giorno["giorno_settimana"]

            for lezione in giorno["lezioni"]:
                ora_str = lezione["ora"].strftime("%H:%M")
                ws.cell(row=row, column=1, value=data_str)
                ws.cell(row=row, column=2, value=giorno_label)
                ws.cell(row=row, column=3, value=ora_str)
                ws.cell(row=row, column=4, value=lezione["materia"])
                ws.cell(row=row, column=5, value=lezione["docente"])
                row += 1

    save_folder = os.path.join(current_app.root_path, "generated_calendars")
    os.makedirs(save_folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"calendario_{timestamp}.xlsx"
    filepath = os.path.join(save_folder, filename)

    wb.save(filepath)

    flash("Calendario generato con successo!", "success")
    return redirect(url_for("orario.lista_versioni"))

# ============================================================
# 2) LISTA VERSIONI (GET)
# ============================================================

@orario_bp.route("/")
@orario_bp.route("/versioni")
def lista_versioni():
    save_folder = os.path.join(current_app.root_path, "generated_calendars")
    os.makedirs(save_folder, exist_ok=True)

    files = sorted(os.listdir(save_folder), reverse=True)

    return render_template("genera_calendario.html", files=files)

# ============================================================
# 3) DOWNLOAD VERSIONE
# ============================================================

@orario_bp.route("/download/<filename>")
def download_calendario(filename):
    folder = os.path.join(current_app.root_path, "generated_calendars")
    return send_from_directory(folder, filename, as_attachment=True)

# ============================================================
# 4) EXPORT XLS (versione singola)
# ============================================================

@orario_bp.route("/export_xls")
def export_xls():
    calendario = genera_calendario_annuale()

    wb = openpyxl.Workbook()
    wb.active.title = "Indice"

    for classe_id, dati in calendario.items():
        ws = wb.create_sheet(title=dati["nome_classe"][:31])

        ws.cell(row=1, column=1, value="Data")
        ws.cell(row=1, column=2, value="Giorno")
        ws.cell(row=1, column=3, value="Ora")
        ws.cell(row=1, column=4, value="Materia")
        ws.cell(row=1, column=5, value="Docente")

        row = 2
        for giorno in dati["calendario"]:
            data_str = giorno["data"].strftime("%d/%m/%Y")
            giorno_label = giorno["giorno_settimana"]

            for lezione in giorno["lezioni"]:
                ora_str = lezione["ora"].strftime("%H:%M")
                ws.cell(row=row, column=1, value=data_str)
                ws.cell(row=row, column=2, value=giorno_label)
                ws.cell(row=row, column=3, value=ora_str)
                ws.cell(row=row, column=4, value=lezione["materia"])
                ws.cell(row=row, column=5, value=lezione["docente"])
                row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="calendario_scolastico.xlsx",
    )

# ============================================================
# 5) REPORT VINCOLI
# ============================================================

@orario_bp.route("/report_vincoli")
def report_vincoli():
    calendario = genera_calendario_annuale()
    violazioni = []

    for classe_id, dati in calendario.items():
        nome_classe = dati["nome_classe"]
        calendario_classe = dati["calendario"]

        materie_classe = MateriaClasse.query.filter_by(classe_id=classe_id).all()
        docente_dict = {d.id: d.nome_docente for d in Docente.query.all()}
        giorni_fissi = GiornoFisso.query.filter_by(classe_id=classe_id).all()

        for giorno in calendario_classe:
            data = giorno["data"]
            giorno_it = giorno["giorno_settimana"]
            lezioni = giorno["lezioni"]

            # Vincoli usando orario_utils
            if orario_utils.classe_in_stage_giorno(classe_id, data):
                for l in lezioni:
                    if l["materia"] not in ("", "STAGE"):
                        violazioni.append(
                            f"{nome_classe} — {data}: lezione '{l['materia']}' durante STAGE"
                        )

            gs = orario_utils.giorno_speciale_classe(classe_id, data)
            if gs:
                for l in lezioni[gs.ore:]:
                    if l["materia"] not in ("", "BLOCCO_SPECIALE"):
                        violazioni.append(
                            f"{nome_classe} — {data}: lezioni oltre il giorno speciale"
                        )

            for gf in giorni_fissi:
                if gf.giorno == giorno_it:
                    count = sum(1 for l in lezioni if l["materia"] == gf.materia.nome)
                    if count < gf.ore:
                        violazioni.append(
                            f"{nome_classe} — {data}: giorno fisso '{gf.materia.nome}' non rispettato"
                        )

            for idx, l in enumerate(lezioni):
                if not l["docente"]:
                    continue

                docente_nome = l["docente"]
                docente = Docente.query.filter_by(nome_docente=docente_nome).first()
                if not docente:
                    continue

                if not orario_utils.docente_disponibile(docente.id, giorno_it, idx):
                    violazioni.append(
                        f"{nome_classe} — {data}: docente {docente_nome} non disponibile alle {l['ora']}"
                    )

    if not violazioni:
        return render_template("report_vincoli.html", report=["TUTTO OK — nessun vincolo violato"])

    return render_template("report_vincoli.html", report=violazioni)

# ============================================================
# 6) DELETE STAGE
# ============================================================

@orario_bp.route("/stage/delete/<int:classe_id>", methods=["POST"])
def delete_stage(classe_id):
    stage = Stage.query.filter_by(classe_id=classe_id).first()
    if stage:
        db.session.delete(stage)
        db.session.commit()

    return redirect(request.referrer or url_for("orario.lista_versioni"))

# ============================================================
# 7) DIAGNOSTICA (VALIDATORE A+B+C)
# ============================================================

@orario_bp.route("/diagnostica")
def diagnostica():
    from app.utils.validator import CLASSI_INFO_CACHE, CALENDARIO_CACHE, stampa_report

    calendario = CALENDARIO_CACHE
    classi_info = CLASSI_INFO_CACHE

    # ⚠️ Controllo corretto: verifica solo None, non dict vuoti
    if calendario is None or classi_info is None:
        return "<h3>Nessun calendario generato.</h3>"

    html = stampa_report(calendario, classi_info)
    return html
